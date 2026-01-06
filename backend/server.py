from fastapi import FastAPI, APIRouter, HTTPException, WebSocket, WebSocketDisconnect
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import httpx
import json
import asyncio
from functools import wraps
import io
import csv
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Google API Configuration
GOOGLE_API_KEY = os.environ.get('GOOGLE_API_KEY', '')
GOOGLE_SHEET_ID = os.environ.get('GOOGLE_SHEET_ID', '')

# Request deduplication cache
ongoing_requests = {}

# Cache configuration
CACHE_TTL_HOURS = int(os.environ.get('CACHE_TTL_HOURS', '24'))  # Default 24 hours

# WebSocket connection management
active_connections: Dict[str, List[WebSocket]] = {
    'supply': [],
    'event': [],
    'contact': []
}

# User presence tracking (initialized after class definitions)
user_presence: Dict[str, Dict[str, Any]] = {
    'supply': {},
    'event': {},
    'contact': {}
}

# Operation history for conflict resolution (initialized after class definitions)
operation_history: Dict[str, List[Any]] = {
    'supply': [],
    'event': [],
    'contact': []
}

# API Metrics
api_metrics = {
    'requests_total': 0,
    'requests_success': 0,
    'requests_error': 0,
    'cache_hits': 0,
    'cache_misses': 0,
    'google_api_calls': 0,
    'google_api_errors': 0
}

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ==================== Utility Functions ====================
def retry_with_backoff(max_retries=3, base_delay=1.0, max_delay=60.0, backoff_factor=2.0):
    """Decorator for retrying API calls with exponential backoff"""
    def decorator(func):
        @wraps(func)
        async def wrapper(*args, **kwargs):
            last_exception = None
            delay = base_delay

            for attempt in range(max_retries + 1):
                try:
                    return await func(*args, **kwargs)
                except Exception as e:
                    last_exception = e

                    # Check if it's a rate limit or server error that should be retried
                    if hasattr(e, 'response') and e.response:
                        status_code = e.response.status_code
                        if status_code in [429, 500, 502, 503, 504]:  # Rate limit or server errors
                            if attempt < max_retries:
                                logger.warning(f"API call failed (attempt {attempt + 1}/{max_retries + 1}): {e}. Retrying in {delay}s...")
                                await asyncio.sleep(delay)
                                delay = min(delay * backoff_factor, max_delay)
                                continue

                    # For other errors or if max retries reached, don't retry
                    break

            logger.error(f"API call failed after {max_retries + 1} attempts: {last_exception}")
            raise last_exception

        return wrapper
    return decorator


def handle_google_api_error(error: Exception, service: str) -> str:
    """Handle and format Google API errors with user-friendly messages"""
    if hasattr(error, 'response') and error.response:
        status_code = error.response.status_code
        if status_code == 403:
            return f"Access denied to {service}. Please check API key permissions."
        elif status_code == 404:
            return f"Resource not found in {service}."
        elif status_code == 429:
            return f"Rate limit exceeded for {service}. Please try again later."
        elif status_code >= 500:
            return f"{service} is temporarily unavailable. Please try again later."
        else:
            return f"Error accessing {service}: {error.response.text[:200]}"
    else:
        return f"Network error accessing {service}: {str(error)}"


def deduplicate_requests(func):
    """Decorator to prevent duplicate concurrent requests"""
    @wraps(func)
    async def wrapper(*args, **kwargs):
        # Create a unique key for this request
        key = f"{func.__name__}:{str(args)}:{str(sorted(kwargs.items()))}"

        # Check if request is already ongoing
        if key in ongoing_requests:
            logger.info(f"Deduplicating request: {key}")
            return await ongoing_requests[key]

        # Create a future for this request
        future = asyncio.Future()
        ongoing_requests[key] = future

        try:
            result = await func(*args, **kwargs)
            future.set_result(result)
            return result
        except Exception as e:
            future.set_exception(e)
            raise
        finally:
            # Clean up after a short delay to allow other waiting requests to complete
            async def cleanup():
                await asyncio.sleep(0.1)
                ongoing_requests.pop(key, None)
            asyncio.create_task(cleanup())

    return wrapper


def track_metrics(func):
    """Decorator to track API call metrics"""
    @wraps(func)
    async def wrapper(*args, **kwargs):
        api_metrics['requests_total'] += 1

        # Check if this is a Google API call
        if func.__name__ in ['fetch_sheet_data', 'fetch_drive_folder', 'fetch_drive_folder_page']:
            api_metrics['google_api_calls'] += 1

        try:
            result = await func(*args, **kwargs)
            api_metrics['requests_success'] += 1
            return result
        except Exception as e:
            api_metrics['requests_error'] += 1
            if func.__name__ in ['fetch_sheet_data', 'fetch_drive_folder', 'fetch_drive_folder_page']:
                api_metrics['google_api_errors'] += 1
            raise

    return wrapper


# ==================== Export Utility Functions ====================
def export_to_csv(data: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> str:
    """Export data to CSV format"""
    if not data:
        return ""

    # Determine columns if not provided
    if not columns:
        columns = list(data[0].keys()) if data else []

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    writer.writerows(data)

    return output.getvalue()


def export_to_excel(data: List[Dict[str, Any]], columns: Optional[List[str]] = None,
                   title: str = "Data Export") -> bytes:
    """Export data to Excel format with styling"""
    if not data:
        # Create empty workbook
        wb = Workbook()
        return io.BytesIO().getvalue()

    # Determine columns if not provided
    if not columns:
        columns = list(data[0].keys()) if data else []

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal="left", vertical="center")

    # Add title
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:' + chr(ord('A') + len(columns) - 1) + '1')

    # Add headers
    for col_num, column in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_num, value=column)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = alignment

    # Add data
    for row_num, row in enumerate(data, 3):
        for col_num, column in enumerate(columns, 1):
            value = row.get(column, '')
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            cell.alignment = alignment

    # Auto-adjust column widths
    for col_num, column in enumerate(columns, 1):
        max_length = len(column)
        for row in data:
            value = str(row.get(column, ''))
            max_length = max(max_length, len(value))
        ws.column_dimensions[chr(ord('A') + col_num - 1)].width = min(max_length + 2, 50)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def export_to_pdf(data: List[Dict[str, Any]], columns: Optional[List[str]] = None,
                 title: str = "Data Export") -> bytes:
    """Export data to PDF format with professional styling"""
    if not data:
        # Create empty PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    # Determine columns if not provided
    if not columns:
        columns = list(data[0].keys()) if data else []

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                          rightMargin=72, leftMargin=72,
                          topMargin=72, bottomMargin=72)

    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    # Prepare table data
    table_data = [columns]  # Header row

    for row in data:
        table_row = []
        for column in columns:
            value = str(row.get(column, ''))
            table_row.append(value)
        table_data.append(table_row)

    # Create table
    table = Table(table_data, repeatRows=1)

    # Table styling
    table_style = TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),

        # Data styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),

        # Grid styling
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BOX', (0, 0), (-1, -1), 1, colors.black),

        # Alternating row colors
        ('BACKGROUND', (0, 1), (-1, 1), colors.lightgrey),
        ('BACKGROUND', (0, 3), (-1, 3), colors.lightgrey),
        ('BACKGROUND', (0, 5), (-1, 5), colors.lightgrey),
    ])

    table.setStyle(table_style)

    # Add metadata
    metadata_style = ParagraphStyle(
        'Metadata',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=2  # Right alignment
    )

    elements.append(table)
    elements.append(Spacer(1, 20))

    # Add footer with export info
    footer_text = f"Generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')} | {len(data)} records"
    elements.append(Paragraph(footer_text, metadata_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


# ==================== Models ====================
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class StatusCheckCreate(BaseModel):
    client_name: str


class CacheData(BaseModel):
    resource_type: str
    resource_id: str
    data: Dict[str, Any]
    cached_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


# ==================== Collaborative Editing Models ====================
class UserPresence(BaseModel):
    user_id: str
    username: str
    color: str
    cursor_position: Optional[Dict[str, Any]] = None
    last_seen: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class CollaborativeOperation(BaseModel):
    operation_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    username: str
    sheet_type: str  # 'supply', 'event', 'contact'
    operation_type: str  # 'insert', 'update', 'delete'
    row_index: Optional[int] = None
    column_key: Optional[str] = None
    old_value: Any = None
    new_value: Any = None
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ConflictResolution(BaseModel):
    operation_id: str
    resolution: str  # 'accept', 'reject', 'merge'
    resolved_by: str
    resolved_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


# ==================== Bulk Operations Models ====================
class BulkOperation(BaseModel):
    operation_id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    username: str
    sheet_type: str
    operation_type: str  # 'bulk_update', 'bulk_insert', 'bulk_delete', 'bulk_import'
    operations: List[CollaborativeOperation]
    total_operations: int
    status: str = 'pending'  # 'pending', 'processing', 'completed', 'failed'
    progress: int = 0  # 0-100
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None
    error_message: Optional[str] = None


class BulkUpdateRequest(BaseModel):
    sheet_type: str
    updates: List[Dict[str, Any]]  # List of {row_index, column_key, new_value}
    description: Optional[str] = None


class BulkDeleteRequest(BaseModel):
    sheet_type: str
    row_indices: List[int]
    description: Optional[str] = None


class BulkImportRequest(BaseModel):
    sheet_type: str
    data: List[Dict[str, Any]]
    update_existing: bool = False  # If true, update existing rows, else append
    description: Optional[str] = None


# ==================== Google Sheets API ====================
def get_sheet_url(sheet_name: str) -> str:
    """Generate Google Sheets API URL"""
    return f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/gviz/tq?tqx=out:json&sheet={sheet_name}"


@retry_with_backoff(max_retries=3, base_delay=1.0, max_delay=30.0)
@deduplicate_requests
@track_metrics
async def fetch_sheet_data(sheet_name: str) -> List[Dict[str, Any]]:
    """Fetch data from Google Sheets with retry logic"""
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(get_sheet_url(sheet_name))
            response.raise_for_status()

            text = response.text
            # Remove the JavaScript wrapper
            json_str = text[47:-2]
            data = json.loads(json_str)

            if 'table' not in data or 'rows' not in data['table']:
                logger.warning(f"No data found in sheet '{sheet_name}'")
                return []

            headers = [col.get('label', '') for col in data['table']['cols']]
            rows = []

            for row in data['table']['rows']:
                obj = {}
                for i, cell in enumerate(row['c']):
                    obj[headers[i]] = cell['v'] if cell else ''
                rows.append(obj)

            logger.info(f"Successfully fetched {len(rows)} rows from sheet '{sheet_name}'")
            return rows
    except Exception as e:
        error_msg = handle_google_api_error(e, "Google Sheets")
        logger.error(f"Error fetching sheet data for '{sheet_name}': {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)


# ==================== Google Drive API ====================
@retry_with_backoff(max_retries=3, base_delay=1.0, max_delay=30.0)
@track_metrics
async def fetch_drive_folder_page(folder_id: str, page_token: str = None) -> Dict[str, Any]:
    """Fetch a page of files and folders from Google Drive"""
    base_url = "https://www.googleapis.com/drive/v3/files"
    params = {
        'q': f"'{folder_id}' in parents",
        'fields': 'nextPageToken,files(id,name,mimeType,thumbnailLink,webViewLink,iconLink,modifiedTime,size)',
        'key': GOOGLE_API_KEY,
        'pageSize': 100,  # Maximum allowed
        'orderBy': 'name'
    }

    if page_token:
        params['pageToken'] = page_token

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(base_url, params=params)
            response.raise_for_status()
            return response.json()
    except Exception as e:
        error_msg = handle_google_api_error(e, "Google Drive")
        logger.error(f"Error fetching Drive folder page for '{folder_id}': {error_msg}")
        raise


@deduplicate_requests
@track_metrics
async def fetch_drive_folder(folder_id: str) -> Dict[str, Any]:
    """Fetch all files and folders from Google Drive with pagination support"""
    try:
        all_files = []
        page_token = None

        while True:
            data = await fetch_drive_folder_page(folder_id, page_token)
            all_files.extend(data.get('files', []))
            page_token = data.get('nextPageToken')

            if not page_token:
                break

        folders = [f for f in all_files if f.get('mimeType') == 'application/vnd.google-apps.folder']
        files = [f for f in all_files if f.get('mimeType') != 'application/vnd.google-apps.folder']

        logger.info(f"Successfully fetched {len(folders)} folders and {len(files)} files from Drive folder '{folder_id}'")
        return {'folders': folders, 'files': files}

    except Exception as e:
        error_msg = handle_google_api_error(e, "Google Drive")
        logger.error(f"Error fetching Drive folder '{folder_id}': {error_msg}")
        raise HTTPException(status_code=500, detail=error_msg)


# ==================== Caching Functions ====================
async def cache_resource(resource_type: str, resource_id: str, data: Any):
    """Cache resource data to MongoDB"""
    try:
        cache_doc = {
            'resource_type': resource_type,
            'resource_id': resource_id,
            'data': data,
            'cached_at': datetime.now(timezone.utc).isoformat()
        }
        
        await db.cache.update_one(
            {'resource_type': resource_type, 'resource_id': resource_id},
            {'$set': cache_doc},
            upsert=True
        )
        logger.info(f"Cached {resource_type}: {resource_id}")
    except Exception as e:
        logger.error(f"Error caching resource: {e}")


async def get_cached_resource(resource_type: str, resource_id: str) -> Optional[Dict[str, Any]]:
    """Get cached resource from MongoDB with expiration check"""
    try:
        cache_doc = await db.cache.find_one(
            {'resource_type': resource_type, 'resource_id': resource_id},
            {'_id': 0}
        )

        if not cache_doc:
            return None

        # Check if cache is expired
        cached_at = cache_doc.get('cached_at')
        if cached_at:
            try:
                # Parse the ISO string to datetime
                if isinstance(cached_at, str):
                    cached_time = datetime.fromisoformat(cached_at.replace('Z', '+00:00'))
                else:
                    cached_time = cached_at

                # Check if cache is expired
                now = datetime.now(timezone.utc)
                cache_age = now - cached_time
                max_age = timedelta(hours=CACHE_TTL_HOURS)

                if cache_age > max_age:
                    logger.info(f"Cache expired for {resource_type}:{resource_id} (age: {cache_age})")
                    # Remove expired cache entry
                    await db.cache.delete_one(
                        {'resource_type': resource_type, 'resource_id': resource_id}
                    )
                    return None

            except (ValueError, TypeError) as e:
                logger.warning(f"Error parsing cache timestamp: {e}")

        return cache_doc
    except Exception as e:
        logger.error(f"Error getting cached resource: {e}")
        return None


# ==================== API Routes ====================

@api_router.get("/")
async def root():
    return {"message": "MDRRMO Dashboard API"}


@api_router.get("/health")
async def health_check():
    return {
        "status": "healthy",
        "timestamp": datetime.now(timezone.utc).isoformat()
    }


# ==================== Google Sheets Routes ====================

@api_router.get("/sheets/{sheet_name}")
async def get_sheet_data(sheet_name: str):
    """Get data from Google Sheets with caching"""
    try:
        # Try to fetch fresh data
        data = await fetch_sheet_data(sheet_name)

        # Cache the data
        await cache_resource('sheet', sheet_name, data)

        return {
            'success': True,
            'data': data,
            'cached': False,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        # If fetch fails, try to return cached data
        logger.warning(f"Failed to fetch sheet data, trying cache: {e}")
        cached = await get_cached_resource('sheet', sheet_name)

        if cached:
            api_metrics['cache_hits'] += 1
            return {
                'success': True,
                'data': cached.get('data', []),
                'cached': True,
                'cached_at': cached.get('cached_at'),
                'timestamp': datetime.now(timezone.utc).isoformat()
            }

        api_metrics['cache_misses'] += 1
        raise HTTPException(status_code=500, detail="Failed to fetch data and no cache available")


@api_router.get("/cache/sheets/{sheet_name}")
async def get_cached_sheet_data(sheet_name: str):
    """Get cached sheet data for offline access"""
    cached = await get_cached_resource('sheet', sheet_name)
    
    if not cached:
        raise HTTPException(status_code=404, detail="No cached data available")
    
    return {
        'success': True,
        'data': cached.get('data', []),
        'cached_at': cached.get('cached_at'),
        'timestamp': datetime.now(timezone.utc).isoformat()
    }


# ==================== Google Drive Routes ====================

@api_router.get("/drive/folder/{folder_id}")
async def get_drive_folder(folder_id: str):
    """Get Google Drive folder contents with caching"""
    try:
        # Try to fetch fresh data
        data = await fetch_drive_folder(folder_id)

        # Cache the data
        await cache_resource('drive_folder', folder_id, data)

        return {
            'success': True,
            'data': data,
            'cached': False,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        # If fetch fails, try to return cached data
        logger.warning(f"Failed to fetch Drive folder, trying cache: {e}")
        cached = await get_cached_resource('drive_folder', folder_id)

        if cached:
            api_metrics['cache_hits'] += 1
            return {
                'success': True,
                'data': cached.get('data', {'folders': [], 'files': []}),
                'cached': True,
                'cached_at': cached.get('cached_at'),
                'timestamp': datetime.now(timezone.utc).isoformat()
            }

        api_metrics['cache_misses'] += 1
        raise HTTPException(status_code=500, detail="Failed to fetch data and no cache available")


@api_router.get("/cache/drive/folder/{folder_id}")
async def get_cached_drive_folder(folder_id: str):
    """Get cached Drive folder data for offline access"""
    cached = await get_cached_resource('drive_folder', folder_id)
    
    if not cached:
        raise HTTPException(status_code=404, detail="No cached data available")
    
    return {
        'success': True,
        'data': cached.get('data', {'folders': [], 'files': []}),
        'cached_at': cached.get('cached_at'),
        'timestamp': datetime.now(timezone.utc).isoformat()
    }


# ==================== Sync Routes ====================

@api_router.post("/sync/all")
async def sync_all_data():
    """Sync all data from Google Sheets and Drive"""
    try:
        results = {
            'sheets': {},
            'drive_folders': {}
        }
        
        # Sync all sheets
        sheet_names = ['supply', 'event', 'contact']
        for sheet_name in sheet_names:
            try:
                data = await fetch_sheet_data(sheet_name)
                await cache_resource('sheet', sheet_name, data)
                results['sheets'][sheet_name] = 'success'
            except Exception as e:
                results['sheets'][sheet_name] = f'failed: {str(e)}'
        
        # Sync critical Drive folders
        drive_folders = {
            'documents': '15_xiFeXu_vdIe2CYrjGaRCAho2OqhGvo',
            'photos': '1O1WlCjMvZ5lVcrOIGNMlBY4ZuQ-zEarg',
            'panorama': '1tsbcsTEfg5RLHLJLYXR41avy9SrajsqM',
            'administrative': '1Wh2wSQuyzHiz25Vbr4ICETj18RRUEpvi',
            'topographic': '1Y01dJR_YJdixvsi_B9Xs7nQaXD31_Yn2',
            'hazards': '16xy_oUAr6sWb3JE9eNrxYJdAMDRKGYLn',
            'denr': '1yQmtrKfKiMOFA933W0emzeGoexMpUDGM',
            'other': '1MI1aO_-gQwsRbSJsfHY2FI4AOz9Jney1'
        }
        
        for folder_name, folder_id in drive_folders.items():
            try:
                data = await fetch_drive_folder(folder_id)
                await cache_resource('drive_folder', folder_id, data)
                results['drive_folders'][folder_name] = 'success'
            except Exception as e:
                results['drive_folders'][folder_name] = f'failed: {str(e)}'
        
        return {
            'success': True,
            'results': results,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Sync failed: {str(e)}")


@api_router.get("/cache/status")
async def get_cache_status():
    """Get cache status and last sync times"""
    try:
        cached_items = await db.cache.find({}, {'_id': 0, 'resource_type': 1, 'resource_id': 1, 'cached_at': 1}).to_list(100)

        return {
            'success': True,
            'cached_items': cached_items,
            'total_cached': len(cached_items),
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get cache status: {str(e)}")


# ==================== Collaborative Editing Functions ====================
async def broadcast_to_sheet(sheet_type: str, message: Dict[str, Any]):
    """Broadcast message to all connected clients for a specific sheet"""
    disconnected = []
    for i, websocket in enumerate(active_connections[sheet_type]):
        try:
            await websocket.send_json(message)
        except Exception as e:
            logger.warning(f"Failed to send message to websocket {i}: {e}")
            disconnected.append(i)

    # Remove disconnected clients
    for i in reversed(disconnected):
        active_connections[sheet_type].pop(i)


def apply_operation_to_data(data: List[Dict[str, Any]], operation: CollaborativeOperation) -> List[Dict[str, Any]]:
    """Apply a collaborative operation to the data"""
    new_data = [row.copy() for row in data]  # Deep copy

    if operation.operation_type == 'update' and operation.row_index is not None and operation.column_key:
        if 0 <= operation.row_index < len(new_data):
            new_data[operation.row_index][operation.column_key] = operation.new_value
    elif operation.operation_type == 'insert' and operation.row_index is not None:
        new_row = {} if operation.new_value is None else operation.new_value
        new_data.insert(operation.row_index, new_row)
    elif operation.operation_type == 'delete' and operation.row_index is not None:
        if 0 <= operation.row_index < len(new_data):
            new_data.pop(operation.row_index)

    return new_data


async def save_operation_to_history(operation: CollaborativeOperation):
    """Save operation to MongoDB for persistence"""
    try:
        doc = operation.model_dump()
        doc['timestamp'] = doc['timestamp'].isoformat()
        await db.collaborative_operations.insert_one(doc)
    except Exception as e:
        logger.error(f"Failed to save operation to history: {e}")


# ==================== Bulk Operations Functions ====================
async def process_bulk_operation(bulk_op: BulkOperation) -> BulkOperation:
    """Process a bulk operation asynchronously"""
    try:
        bulk_op.status = 'processing'

        # Update progress as we process operations
        total_ops = len(bulk_op.operations)
        completed = 0

        for operation in bulk_op.operations:
            try:
                # Apply the operation (in a real implementation, this would sync with Google Sheets)
                # For now, we'll just save to history
                await save_operation_to_history(operation)

                # Broadcast to connected users
                await broadcast_to_sheet(bulk_op.sheet_type, {
                    'type': 'bulk_operation_progress',
                    'bulk_operation_id': bulk_op.operation_id,
                    'progress': int((completed + 1) / total_ops * 100),
                    'current_operation': completed + 1,
                    'total_operations': total_ops
                })

                completed += 1
                bulk_op.progress = int(completed / total_ops * 100)

            except Exception as e:
                logger.error(f"Failed to process operation {completed + 1}: {e}")
                bulk_op.error_message = f"Failed at operation {completed + 1}: {str(e)}"
                bulk_op.status = 'failed'
                break

        if bulk_op.status != 'failed':
            bulk_op.status = 'completed'
            bulk_op.completed_at = datetime.now(timezone.utc)

            # Final broadcast
            await broadcast_to_sheet(bulk_op.sheet_type, {
                'type': 'bulk_operation_complete',
                'bulk_operation_id': bulk_op.operation_id,
                'total_operations': total_ops
            })

        # Save bulk operation status
        await save_bulk_operation(bulk_op)

        return bulk_op

    except Exception as e:
        bulk_op.status = 'failed'
        bulk_op.error_message = str(e)
        await save_bulk_operation(bulk_op)
        raise


async def save_bulk_operation(bulk_op: BulkOperation):
    """Save bulk operation to database"""
    try:
        doc = bulk_op.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        if doc['completed_at']:
            doc['completed_at'] = doc['completed_at'].isoformat()

        await db.bulk_operations.update_one(
            {'operation_id': bulk_op.operation_id},
            {'$set': doc},
            upsert=True
        )
    except Exception as e:
        logger.error(f"Failed to save bulk operation: {e}")


async def get_bulk_operation_status(operation_id: str) -> Optional[BulkOperation]:
    """Get bulk operation status from database"""
    try:
        doc = await db.bulk_operations.find_one({'operation_id': operation_id})
        if doc:
            # Convert back to datetime objects
            if 'created_at' in doc:
                doc['created_at'] = datetime.fromisoformat(doc['created_at'])
            if 'completed_at' in doc and doc['completed_at']:
                doc['completed_at'] = datetime.fromisoformat(doc['completed_at'])
            return BulkOperation(**doc)
        return None
    except Exception as e:
        logger.error(f"Failed to get bulk operation status: {e}")
        return None


# ==================== WebSocket Endpoints ====================
@api_router.websocket("/ws/collaborate/{sheet_type}/{user_id}")
async def collaborate_websocket(websocket: WebSocket, sheet_type: str, user_id: str):
    """WebSocket endpoint for collaborative editing"""
    if sheet_type not in ['supply', 'event', 'contact']:
        await websocket.close(code=1003)  # Unsupported data
        return

    await websocket.accept()

    # Generate user info
    username = f"User {user_id[:8]}"  # Simple username generation
    user_color = f"#{hash(user_id) % 0xFFFFFF:06x}"  # Simple color generation

    # Add user presence
    presence = UserPresence(
        user_id=user_id,
        username=username,
        color=user_color
    )
    user_presence[sheet_type][user_id] = presence
    active_connections[sheet_type].append(websocket)

    try:
        # Send current user presence to the new user
        await websocket.send_json({
            'type': 'presence_update',
            'users': list(user_presence[sheet_type].values())
        })

        # Broadcast new user presence to others
        await broadcast_to_sheet(sheet_type, {
            'type': 'user_joined',
            'user': presence.model_dump()
        })

        while True:
            try:
                data = await websocket.receive_json()

                if data['type'] == 'operation':
                    operation = CollaborativeOperation(**data['operation'])

                    # Apply operation to in-memory data (in a real app, you'd sync with Google Sheets)
                    # For now, we'll just broadcast the operation
                    operation_history[sheet_type].append(operation)

                    # Save to database
                    await save_operation_to_history(operation)

                    # Broadcast to all clients except sender
                    for ws in active_connections[sheet_type]:
                        if ws != websocket:
                            await ws.send_json({
                                'type': 'operation',
                                'operation': operation.model_dump()
                            })

                elif data['type'] == 'cursor_update':
                    # Update cursor position
                    if user_id in user_presence[sheet_type]:
                        user_presence[sheet_type][user_id].cursor_position = data.get('position')
                        user_presence[sheet_type][user_id].last_seen = datetime.now(timezone.utc)

                        # Broadcast cursor update
                        await broadcast_to_sheet(sheet_type, {
                            'type': 'cursor_update',
                            'user_id': user_id,
                            'position': data.get('position')
                        })

            except json.JSONDecodeError:
                await websocket.send_json({'type': 'error', 'message': 'Invalid JSON'})

    except WebSocketDisconnect:
        logger.info(f"WebSocket disconnected for user {user_id} on {sheet_type}")

    finally:
        # Remove user presence
        if user_id in user_presence[sheet_type]:
            del user_presence[sheet_type][user_id]

        # Remove from active connections
        if websocket in active_connections[sheet_type]:
            active_connections[sheet_type].remove(websocket)

        # Broadcast user left
        await broadcast_to_sheet(sheet_type, {
            'type': 'user_left',
            'user_id': user_id
        })


# ==================== Bulk Operations API Endpoints ====================
@api_router.post("/bulk/update")
async def bulk_update(request: BulkUpdateRequest, user_id: str = "system"):
    """Perform bulk update operations"""
    try:
        # Create individual operations
        operations = []
        for update in request.updates:
            operation = CollaborativeOperation(
                user_id=user_id,
                username=f"User {user_id[:8]}",
                sheet_type=request.sheet_type,
                operation_type='update',
                row_index=update['row_index'],
                column_key=update['column_key'],
                old_value=update.get('old_value'),
                new_value=update['new_value']
            )
            operations.append(operation)

        # Create bulk operation
        bulk_op = BulkOperation(
            user_id=user_id,
            username=f"User {user_id[:8]}",
            sheet_type=request.sheet_type,
            operation_type='bulk_update',
            operations=operations,
            total_operations=len(operations)
        )

        # Start processing in background
        import asyncio
        asyncio.create_task(process_bulk_operation(bulk_op))

        return {
            'success': True,
            'bulk_operation_id': bulk_op.operation_id,
            'total_operations': len(operations),
            'message': f'Bulk update started for {len(operations)} operations'
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Bulk update failed: {str(e)}")


@api_router.post("/bulk/delete")
async def bulk_delete(request: BulkDeleteRequest, user_id: str = "system"):
    """Perform bulk delete operations"""
    try:
        # Create individual operations (sorted in reverse order to maintain indices)
        operations = []
        for row_index in sorted(request.row_indices, reverse=True):
            operation = CollaborativeOperation(
                user_id=user_id,
                username=f"User {user_id[:8]}",
                sheet_type=request.sheet_type,
                operation_type='delete',
                row_index=row_index
            )
            operations.append(operation)

        # Create bulk operation
        bulk_op = BulkOperation(
            user_id=user_id,
            username=f"User {user_id[:8]}",
            sheet_type=request.sheet_type,
            operation_type='bulk_delete',
            operations=operations,
            total_operations=len(operations)
        )

        # Start processing in background
        import asyncio
        asyncio.create_task(process_bulk_operation(bulk_op))

        return {
            'success': True,
            'bulk_operation_id': bulk_op.operation_id,
            'total_operations': len(operations),
            'message': f'Bulk delete started for {len(operations)} rows'
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Bulk delete failed: {str(e)}")


@api_router.post("/bulk/import")
async def bulk_import(request: BulkImportRequest, user_id: str = "system"):
    """Perform bulk import operations"""
    try:
        operations = []

        if request.update_existing:
            # Update existing rows
            for i, row_data in enumerate(request.data):
                for column_key, new_value in row_data.items():
                    operation = CollaborativeOperation(
                        user_id=user_id,
                        username=f"User {user_id[:8]}",
                        sheet_type=request.sheet_type,
                        operation_type='update',
                        row_index=i,
                        column_key=column_key,
                        new_value=new_value
                    )
                    operations.append(operation)
        else:
            # Insert new rows
            for row_data in request.data:
                operation = CollaborativeOperation(
                    user_id=user_id,
                    username=f"User {user_id[:8]}",
                    sheet_type=request.sheet_type,
                    operation_type='insert',
                    new_value=row_data
                )
                operations.append(operation)

        # Create bulk operation
        bulk_op = BulkOperation(
            user_id=user_id,
            username=f"User {user_id[:8]}",
            sheet_type=request.sheet_type,
            operation_type='bulk_import',
            operations=operations,
            total_operations=len(operations)
        )

        # Start processing in background
        import asyncio
        asyncio.create_task(process_bulk_operation(bulk_op))

        return {
            'success': True,
            'bulk_operation_id': bulk_op.operation_id,
            'total_operations': len(operations),
            'message': f'Bulk import started for {len(request.data)} rows'
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Bulk import failed: {str(e)}")


@api_router.get("/bulk/status/{operation_id}")
async def get_bulk_status(operation_id: str):
    """Get status of a bulk operation"""
    try:
        bulk_op = await get_bulk_operation_status(operation_id)
        if not bulk_op:
            raise HTTPException(status_code=404, detail="Bulk operation not found")

        return {
            'success': True,
            'operation_id': bulk_op.operation_id,
            'status': bulk_op.status,
            'progress': bulk_op.progress,
            'total_operations': bulk_op.total_operations,
            'created_at': bulk_op.created_at.isoformat(),
            'completed_at': bulk_op.completed_at.isoformat() if bulk_op.completed_at else None,
            'error_message': bulk_op.error_message
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get bulk status: {str(e)}")


# ==================== Export API Endpoints ====================
@api_router.get("/export/{sheet_type}/{format}")
async def export_data(sheet_type: str, format: str, columns: Optional[str] = None):
    """Export data in various formats (csv, excel, pdf)"""
    if format not in ['csv', 'excel', 'pdf']:
        raise HTTPException(status_code=400, detail="Invalid format. Supported: csv, excel, pdf")

    if sheet_type not in ['supply', 'event', 'contact']:
        raise HTTPException(status_code=400, detail="Invalid sheet type")

    try:
        # Get data from cache or fresh fetch
        cached = await get_cached_resource('sheet', sheet_type)
        if cached and cached.get('data'):
            data = cached['data']
        else:
            # Fetch fresh data
            result = await fetch_sheet_data(sheet_type)
            data = result

        if not data:
            raise HTTPException(status_code=404, detail="No data available for export")

        # Parse columns parameter
        column_list = None
        if columns:
            try:
                column_list = [col.strip() for col in columns.split(',') if col.strip()]
            except:
                pass

        # Generate title
        titles = {
            'supply': 'Supply Inventory',
            'event': 'Calendar Events',
            'contact': 'Contact Directory'
        }
        title = titles.get(sheet_type, 'Data Export')

        # Export based on format
        if format == 'csv':
            content = export_to_csv(data, column_list)
            media_type = 'text/csv'
            filename = f"{sheet_type}_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"

        elif format == 'excel':
            content = export_to_excel(data, column_list, title)
            media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            filename = f"{sheet_type}_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

        elif format == 'pdf':
            content = export_to_pdf(data, column_list, title)
            media_type = 'application/pdf'
            filename = f"{sheet_type}_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf"

        from fastapi.responses import StreamingResponse
        import io

        if isinstance(content, str):
            # CSV content is a string
            content_bytes = content.encode('utf-8')
        else:
            # Excel/PDF content is already bytes
            content_bytes = content

        def iter_content():
            yield content_bytes

        return StreamingResponse(
            iter_content(),
            media_type=media_type,
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Length': str(len(content_bytes))
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Export failed for {sheet_type} in {format}: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@api_router.get("/metrics")
async def get_api_metrics():
    """Get API performance metrics"""
    try:
        # Calculate rates
        total_requests = api_metrics['requests_total']
        success_rate = (api_metrics['requests_success'] / total_requests * 100) if total_requests > 0 else 0
        cache_hit_rate = (api_metrics['cache_hits'] / (api_metrics['cache_hits'] + api_metrics['cache_misses']) * 100) if (api_metrics['cache_hits'] + api_metrics['cache_misses']) > 0 else 0

        return {
            'success': True,
            'metrics': {
                **api_metrics,
                'success_rate_percent': round(success_rate, 2),
                'cache_hit_rate_percent': round(cache_hit_rate, 2)
            },
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get metrics: {str(e)}")


# ==================== Status Check Routes (Original) ====================

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj


@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks


# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
